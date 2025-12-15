# -*- coding: utf-8 -*-

"""
Excel-отчёт для Artbazar AI

Назначение:
- дать пользователю РЕАЛЬНУЮ таблицу
- данные можно фильтровать, считать, дополнять
- Excel = инструмент, а не витрина
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def build_excel_report(history: list) -> io.BytesIO:
    """
    history: список словарей из context.user_data["history"]

    Каждый элемент ожидаемо содержит:
    - type
    - date
    - summary
    - risk_level (если есть)
    - demand_type / seasonality / competition / resource (если есть)
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Artbazar Report"

    headers = [
        "Дата",
        "Тип анализа",
        "Краткий итог",
        "Уровень риска",
        "Тип спроса",
        "Сезонность",
        "Конкуренция",
        "Ресурс",
    ]

    ws.append(headers)

    for item in history:
        ws.append([
            item.get("date", ""),
            item.get("type", ""),
            item.get("summary", ""),
            item.get("risk_level", ""),
            item.get("demand_type", ""),
            item.get("seasonality", ""),
            item.get("competition", ""),
            item.get("resource", ""),
        ])

    # Автоширина колонок
    for col_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)

    # Метаданные
    meta = wb.create_sheet("Meta")
    meta.append(["Сформировано"])
    meta.append([datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    meta.append([])
    meta.append(["Artbazar AI — Excel отчёт"])
    meta.append(["Данные предназначены для самостоятельного анализа."])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
